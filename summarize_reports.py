"""
summarize_reports.py
--------------------
Reads one or more Checkmarx reports (.csv or .xlsx) and prints a
per-file summary:

    filename.csv
    ├─ SQL Injection              :  12
    ├─ Path Traversal             :   6
    └─ HttpOnlyCookies            :   5

Usage:
    # Summarise every file in a folder
    python3 summarize_reports.py --input-dir /path/to/folder

    # Summarise a single file
    python3 summarize_reports.py --input checkmarx.csv

    # Auto-detect (looks for checkmarx.csv / checkmarx.xlsx in current dir)
    python3 summarize_reports.py
"""

import argparse
import csv
import glob
import os
import sys

try:
    import openpyxl
except ImportError:
    openpyxl = None  # XLSX support disabled if not installed

SUPPORTED_EXTS = (".csv", ".xlsx", ".xls")

# ─────────────────────────────────────────────────────────────────────────────
# Readers
# ─────────────────────────────────────────────────────────────────────────────

def _sniff_csv(path):
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with open(path, "r", encoding=enc, newline="") as f:
                sample = f.read(4096)
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
            return enc, dialect.delimiter
        except Exception:
            continue
    return "utf-8-sig", ","


def read_queries_csv(path):
    """Return list of Query values from a CSV file."""
    encoding, delimiter = _sniff_csv(path)
    queries = []
    with open(path, "r", encoding=encoding, newline="") as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        headers = [h.strip() for h in (reader.fieldnames or [])]

        # Find the Query column (case-insensitive)
        query_col = None
        for h in headers:
            if h.lower() == "query":
                query_col = h
                break

        if query_col is None:
            print(f"  [!] 'Query' column not found — available: {headers[:10]}")
            return []

        for row in reader:
            val = row.get(query_col, "").strip()
            if val:
                queries.append(val)
    return queries


def read_queries_xlsx(path):
    """Return list of Query values from an XLSX file."""
    if openpyxl is None:
        sys.exit("[ERROR] openpyxl is not installed. Install it with: pip install openpyxl")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # Find Query column index from header row 1
    query_col = None
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val and str(val).strip().lower() == "query":
            query_col = col
            break

    if query_col is None:
        print("  [!] 'Query' column not found in sheet.")
        return []

    queries = []
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=query_col).value
        if val:
            queries.append(str(val).strip())
    return queries


def read_queries(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        return read_queries_csv(path)
    elif ext in (".xlsx", ".xls"):
        return read_queries_xlsx(path)
    else:
        # Try CSV first, fallback to XLSX
        try:
            return read_queries_csv(path)
        except Exception:
            return read_queries_xlsx(path)


# ─────────────────────────────────────────────────────────────────────────────
# Summary printer
# ─────────────────────────────────────────────────────────────────────────────

def summarise_file(path):
    """Print the vulnerability summary for a single file."""
    from collections import Counter

    filename = os.path.basename(path)
    queries  = read_queries(path)

    if not queries:
        print(f"\n  {filename}")
        print("  └─ (no vulnerability data found)\n")
        return

    counts = Counter(queries)
    total  = sum(counts.values())
    items  = counts.most_common()   # sorted by count descending

    # ── Header ──────────────────────────────────────────────────────────────
    print(f"\n{'─' * 60}")
    print(f"  {filename}")
    print(f"  Total raw findings : {total}   |   Unique vuln types : {len(items)}")
    print(f"{'─' * 60}")

    # ── Rows ────────────────────────────────────────────────────────────────
    max_name_len = max(len(q) for q, _ in items)
    col_w = min(max_name_len, 50)   # cap at 50 chars

    for i, (vuln, count) in enumerate(items):
        branch = "└─" if i == len(items) - 1 else "├─"
        bar    = "█" * min(count, 30)   # mini bar chart (max 30 chars)
        print(f"  {branch} {vuln:<{col_w}}  {count:>4}   {bar}")

    print()


# ─────────────────────────────────────────────────────────────────────────────
# Batch mode
# ─────────────────────────────────────────────────────────────────────────────

def collect_files(folder):
    files = []
    for ext in SUPPORTED_EXTS:
        files.extend(sorted(glob.glob(os.path.join(folder, f"*{ext}"))))
    # Deduplicate
    seen, unique = set(), []
    for f in files:
        key = os.path.normcase(os.path.abspath(f))
        if key not in seen:
            seen.add(key)
            unique.append(f)
    return unique


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Print a vulnerability count summary for Checkmarx report file(s).",
        formatter_class=argparse.RawTextHelpFormatter
    )
    parser.add_argument(
        "--input", "-i", default=None, metavar="FILE",
        help="Single Checkmarx report file (.csv or .xlsx)."
    )
    parser.add_argument(
        "--input-dir", "-I", default=None, metavar="DIR",
        help="Folder of Checkmarx report files — all .csv/.xlsx files are summarised."
    )
    args = parser.parse_args()

    print("\n" + "=" * 60)
    print("  CHECKMARX REPORT SUMMARY")
    print("=" * 60)

    # ── Folder mode ──────────────────────────────────────────────────────────
    if args.input_dir:
        if not os.path.isdir(args.input_dir):
            sys.exit(f"[ERROR] Folder not found: {args.input_dir}")
        files = collect_files(args.input_dir)
        if not files:
            sys.exit(f"[ERROR] No CSV/XLSX files found in: {args.input_dir}")
        for f in files:
            summarise_file(f)
        print(f"  {len(files)} file(s) summarised.\n")
        return

    # ── Single file mode ─────────────────────────────────────────────────────
    if args.input is None:
        for candidate in ("checkmarx.csv", "checkmarx.xlsx", "checkmarx.xls"):
            if os.path.isfile(candidate):
                args.input = candidate
                break
        if args.input is None:
            sys.exit(
                "[ERROR] No input file found.\n"
                "        Use --input <file>  or  --input-dir <folder>"
            )

    if not os.path.isfile(args.input):
        sys.exit(f"[ERROR] File not found: {args.input}")

    summarise_file(args.input)


if __name__ == "__main__":
    main()
